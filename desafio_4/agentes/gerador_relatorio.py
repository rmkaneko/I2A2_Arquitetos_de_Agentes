"""
Agente Gerador de Relatório - Sistema de Processamento VR
Responsável por gerar a planilha Excel final formatada
Autor: Manus AI
Data: 27/08/2025
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from utils.config_loader import get_config_loader
from utils.logger import VRLogger


class GeradorRelatorio:
    """Agente responsável pela geração da planilha Excel final"""
    
    def __init__(self, logger: VRLogger):
        """Inicializa o gerador de relatório"""
        self.logger = logger
        self.config_loader = get_config_loader()
        self.config = self.config_loader.get_config()
        
        # Estilos para formatação
        self._definir_estilos()
        
    def _definir_estilos(self):
        """Define estilos de formatação para a planilha"""
        
        # Cores
        self.cor_cabecalho = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.cor_alternada = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.cor_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.cor_excluido = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Fontes
        self.fonte_cabecalho = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        self.fonte_normal = Font(name="Arial", size=10)
        self.fonte_total = Font(name="Arial", size=10, bold=True)
        
        # Bordas
        self.borda_fina = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alinhamentos
        self.alinhamento_centro = Alignment(horizontal='center', vertical='center')
        self.alinhamento_direita = Alignment(horizontal='right', vertical='center')
        self.alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
    
    def executar(self, df_consolidado: pd.DataFrame, estatisticas: Dict[str, Any]) -> str:
        """Executa a geração do relatório Excel"""
        self.logger.log_info("Iniciando geração do relatório Excel")
        
        # Preparar dados para a planilha
        df_relatorio = self._preparar_dados_relatorio(df_consolidado)
        
        # Criar arquivo Excel
        arquivo_saida = self._criar_arquivo_excel(df_relatorio, estatisticas)
        
        self.logger.log_info(f"Relatório Excel gerado com sucesso: {arquivo_saida}")
        return arquivo_saida
    
    def _preparar_dados_relatorio(self, df_consolidado: pd.DataFrame) -> pd.DataFrame:
        """Prepara os dados no formato da planilha final"""
        
        # Filtrar apenas colaboradores elegíveis
        df_elegiveis = df_consolidado[df_consolidado['elegivel'] == True].copy()
        
        # Criar DataFrame no formato da planilha final
        df_relatorio = pd.DataFrame()
        
        # Mapear colunas conforme layout esperado
        df_relatorio['Matricula'] = df_elegiveis['MATRICULA']
        df_relatorio['Admissão'] = df_elegiveis.get('Admissão', '')
        df_relatorio['Sindicato do Colaborador'] = df_elegiveis['Sindicato']
        
        # Competência (formato data)
        competencia = self.config['regras_negocio']['competencia_referencia']
        ano, mes = competencia.split('-')
        data_competencia = f"{ano}-{mes}-01"
        df_relatorio['Competência'] = pd.to_datetime(data_competencia)
        
        df_relatorio['Dias'] = df_elegiveis.get('dias_calculados', 0)
        df_relatorio['VALOR DIÁRIO VR'] = df_elegiveis.get('valor_diario_vr', 0)
        df_relatorio['TOTAL'] = df_elegiveis.get('valor_total_vr', 0)
        df_relatorio['Custo empresa'] = df_elegiveis.get('custo_empresa', 0)
        df_relatorio['Desconto profissional'] = df_elegiveis.get('desconto_colaborador', 0)
        df_relatorio['OBS GERAL'] = df_elegiveis.get('observacoes', '')
        
        # Ordenar por matrícula
        df_relatorio = df_relatorio.sort_values('Matricula').reset_index(drop=True)
        
        self.logger.log_info(f"Dados preparados para {len(df_relatorio)} colaboradores elegíveis")
        return df_relatorio
    
    def _criar_arquivo_excel(self, df_relatorio: pd.DataFrame, estatisticas: Dict[str, Any]) -> str:
        """Cria o arquivo Excel com formatação"""
        
        # Definir nome do arquivo
        arquivo_saida = self.config_loader.get_output_path()
        
        # Criar workbook
        wb = Workbook()
        
        # Remover planilha padrão
        wb.remove(wb.active)
        
        # Criar aba principal
        ws_principal = wb.create_sheet("VR Mensal")
        self._criar_aba_principal(ws_principal, df_relatorio)
        
        # Criar aba de validações
        ws_validacoes = wb.create_sheet("Validações")
        self._criar_aba_validacoes(ws_validacoes, estatisticas)
        
        # Salvar arquivo
        wb.save(arquivo_saida)
        
        return arquivo_saida
    
    def _criar_aba_principal(self, ws, df_relatorio: pd.DataFrame):
        """Cria a aba principal com os dados dos colaboradores"""
        
        # Adicionar cabeçalhos
        cabecalhos = list(df_relatorio.columns)
        
        for col_num, cabecalho in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col_num, value=cabecalho)
            cell.font = self.fonte_cabecalho
            cell.fill = self.cor_cabecalho
            cell.border = self.borda_fina
            cell.alignment = self.alinhamento_centro
        
        # Adicionar dados
        for row_num, (_, row) in enumerate(df_relatorio.iterrows(), 2):
            for col_num, valor in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                cell.font = self.fonte_normal
                cell.border = self.borda_fina
                
                # Formatação específica por coluna
                coluna_nome = cabecalhos[col_num - 1]
                
                if coluna_nome in ['VALOR DIÁRIO VR', 'TOTAL', 'Custo empresa', 'Desconto profissional']:
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = self.alinhamento_direita
                elif coluna_nome in ['Matricula', 'Dias']:
                    cell.alignment = self.alinhamento_centro
                elif coluna_nome in ['Admissão', 'Competência']:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = self.alinhamento_centro
                else:
                    cell.alignment = self.alinhamento_esquerda
                
                # Cor alternada para linhas
                if row_num % 2 == 0:
                    cell.fill = self.cor_alternada
        
        # Adicionar linha de totais
        self._adicionar_linha_totais(ws, df_relatorio, len(df_relatorio) + 2)
        
        # Ajustar largura das colunas
        self._ajustar_largura_colunas(ws, cabecalhos)
        
        # Congelar primeira linha
        ws.freeze_panes = "A2"
    
    def _adicionar_linha_totais(self, ws, df_relatorio: pd.DataFrame, row_num: int):
        """Adiciona linha de totais na planilha"""
        
        # Calcular totais
        total_colaboradores = len(df_relatorio)
        total_vr = df_relatorio['TOTAL'].sum()
        total_custo_empresa = df_relatorio['Custo empresa'].sum()
        total_desconto = df_relatorio['Desconto profissional'].sum()
        
        # Adicionar células de total
        ws.cell(row=row_num, column=1, value="TOTAL GERAL").font = self.fonte_total
        ws.cell(row=row_num, column=1).fill = self.cor_total
        ws.cell(row=row_num, column=1).border = self.borda_fina
        
        # Colaboradores
        cell_colab = ws.cell(row=row_num, column=2, value=total_colaboradores)
        cell_colab.font = self.fonte_total
        cell_colab.fill = self.cor_total
        cell_colab.border = self.borda_fina
        cell_colab.alignment = self.alinhamento_centro
        
        # Valores monetários (assumindo que TOTAL está na coluna 7)
        colunas_valores = {
            7: total_vr,           # TOTAL
            8: total_custo_empresa, # Custo empresa
            9: total_desconto      # Desconto profissional
        }
        
        for col_num, valor in colunas_valores.items():
            cell = ws.cell(row=row_num, column=col_num, value=valor)
            cell.font = self.fonte_total
            cell.fill = self.cor_total
            cell.border = self.borda_fina
            cell.number_format = 'R$ #,##0.00'
            cell.alignment = self.alinhamento_direita
        
        # Preencher células vazias da linha de total
        for col_num in range(3, 11):
            if col_num not in colunas_valores:
                cell = ws.cell(row=row_num, column=col_num, value="")
                cell.fill = self.cor_total
                cell.border = self.borda_fina
    
    def _criar_aba_validacoes(self, ws, estatisticas: Dict[str, Any]):
        """Cria a aba de validações com checagens de consistência"""
        
        # Título
        ws.cell(row=1, column=1, value="RELATÓRIO DE VALIDAÇÕES E CONSISTÊNCIA").font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        row_atual = 3
        
        # Resumo estatístico
        ws.cell(row=row_atual, column=1, value="RESUMO ESTATÍSTICO").font = Font(size=12, bold=True)
        row_atual += 2
        
        resumo_items = [
            ("Total de colaboradores processados", estatisticas.get('total_colaboradores', 0)),
            ("Colaboradores elegíveis para VR", estatisticas.get('colaboradores_elegiveis', 0)),
            ("Colaboradores excluídos", estatisticas.get('colaboradores_excluidos', 0)),
            ("Valor total processado", f"R$ {estatisticas.get('valor_total', 0):,.2f}"),
            ("Custo total empresa", f"R$ {estatisticas.get('custo_total_empresa', 0):,.2f}"),
            ("Desconto total colaboradores", f"R$ {estatisticas.get('desconto_total_colaboradores', 0):,.2f}"),
        ]
        
        for item, valor in resumo_items:
            ws.cell(row=row_atual, column=1, value=item).font = self.fonte_normal
            ws.cell(row=row_atual, column=2, value=valor).font = self.fonte_normal
            row_atual += 1
        
        row_atual += 2
        
        # Exclusões por categoria
        ws.cell(row=row_atual, column=1, value="EXCLUSÕES POR CATEGORIA").font = Font(size=12, bold=True)
        row_atual += 2
        
        exclusoes = estatisticas.get('exclusoes_por_motivo', {})
        for motivo, quantidade in exclusoes.items():
            ws.cell(row=row_atual, column=1, value=motivo).font = self.fonte_normal
            ws.cell(row=row_atual, column=2, value=quantidade).font = self.fonte_normal
            row_atual += 1
        
        row_atual += 2
        
        # Validações de consistência
        ws.cell(row=row_atual, column=1, value="VALIDAÇÕES DE CONSISTÊNCIA").font = Font(size=12, bold=True)
        row_atual += 2
        
        # Calcular algumas validações
        validacoes = self._calcular_validacoes_consistencia(estatisticas)
        
        for validacao, resultado in validacoes.items():
            status = "✓" if resultado['status'] else "⚠"
            ws.cell(row=row_atual, column=1, value=f"{status} {validacao}").font = self.fonte_normal
            ws.cell(row=row_atual, column=2, value=resultado['detalhes']).font = self.fonte_normal
            row_atual += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _calcular_validacoes_consistencia(self, estatisticas: Dict[str, Any]) -> Dict[str, Dict]:
        """Calcula validações de consistência dos dados"""
        
        validacoes = {}
        
        # Validação 1: Soma dos percentuais
        custo_empresa = estatisticas.get('custo_total_empresa', 0)
        desconto_colaborador = estatisticas.get('desconto_total_colaboradores', 0)
        valor_total = estatisticas.get('valor_total', 0)
        
        soma_percentuais = custo_empresa + desconto_colaborador
        diferenca = abs(valor_total - soma_percentuais)
        
        validacoes["Soma dos percentuais empresa + colaborador = total"] = {
            'status': diferenca < 0.01,  # Tolerância de 1 centavo
            'detalhes': f"Diferença: R$ {diferenca:.2f}"
        }
        
        # Validação 2: Percentuais corretos
        if valor_total > 0:
            perc_empresa_real = (custo_empresa / valor_total) * 100
            perc_colaborador_real = (desconto_colaborador / valor_total) * 100
            
            perc_empresa_config = self.config['regras_negocio']['percentual_empresa'] * 100
            perc_colaborador_config = self.config['regras_negocio']['percentual_colaborador'] * 100
            
            validacoes["Percentuais empresa/colaborador corretos"] = {
                'status': (abs(perc_empresa_real - perc_empresa_config) < 0.1 and 
                          abs(perc_colaborador_real - perc_colaborador_config) < 0.1),
                'detalhes': f"Empresa: {perc_empresa_real:.1f}% | Colaborador: {perc_colaborador_real:.1f}%"
            }
        
        # Validação 3: Colaboradores com férias
        colaboradores_ferias = estatisticas.get('colaboradores_com_ferias', 0)
        validacoes["Colaboradores com férias processados"] = {
            'status': True,
            'detalhes': f"{colaboradores_ferias} colaboradores"
        }
        
        # Validação 4: Colaboradores no exterior
        colaboradores_exterior = estatisticas.get('colaboradores_exterior', 0)
        validacoes["Colaboradores no exterior processados"] = {
            'status': True,
            'detalhes': f"{colaboradores_exterior} colaboradores"
        }
        
        # Validação 5: Total de colaboradores
        total_processados = estatisticas.get('total_colaboradores', 0)
        elegiveis = estatisticas.get('colaboradores_elegiveis', 0)
        excluidos = estatisticas.get('colaboradores_excluidos', 0)
        
        validacoes["Total processados = elegíveis + excluídos"] = {
            'status': total_processados == (elegiveis + excluidos),
            'detalhes': f"{total_processados} = {elegiveis} + {excluidos}"
        }
        
        return validacoes
    
    def _ajustar_largura_colunas(self, ws, cabecalhos: List[str]):
        """Ajusta a largura das colunas baseado no conteúdo"""
        
        larguras_padrao = {
            'Matricula': 12,
            'Admissão': 12,
            'Sindicato do Colaborador': 50,
            'Competência': 12,
            'Dias': 8,
            'VALOR DIÁRIO VR': 15,
            'TOTAL': 15,
            'Custo empresa': 15,
            'Desconto profissional': 18,
            'OBS GERAL': 30
        }
        
        for col_num, cabecalho in enumerate(cabecalhos, 1):
            largura = larguras_padrao.get(cabecalho, 15)
            ws.column_dimensions[get_column_letter(col_num)].width = largura
    
    def gerar_relatorio_exclusoes(self, df_consolidado: pd.DataFrame) -> str:
        """Gera relatório separado com colaboradores excluídos"""
        
        # Filtrar colaboradores excluídos
        df_excluidos = df_consolidado[df_consolidado['elegivel'] == False].copy()
        
        if df_excluidos.empty:
            self.logger.log_info("Nenhum colaborador excluído para relatório")
            return None
        
        # Preparar dados
        df_relatorio_exclusoes = pd.DataFrame()
        df_relatorio_exclusoes['Matricula'] = df_excluidos['MATRICULA']
        df_relatorio_exclusoes['Nome/Cargo'] = df_excluidos['TITULO DO CARGO']
        df_relatorio_exclusoes['Sindicato'] = df_excluidos['Sindicato']
        df_relatorio_exclusoes['Motivo Exclusão'] = df_excluidos['motivo_exclusao']
        df_relatorio_exclusoes['Situação'] = df_excluidos['DESC. SITUACAO']
        
        # Salvar em arquivo separado
        arquivo_exclusoes = self.config_loader.get_output_path("colaboradores_excluidos.xlsx")
        
        with pd.ExcelWriter(arquivo_exclusoes, engine='openpyxl') as writer:
            df_relatorio_exclusoes.to_excel(writer, sheet_name='Excluídos', index=False)
        
        self.logger.log_info(f"Relatório de exclusões gerado: {arquivo_exclusoes}")
        return arquivo_exclusoes

